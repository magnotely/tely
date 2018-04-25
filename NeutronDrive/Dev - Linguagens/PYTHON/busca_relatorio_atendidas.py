# import time
import glob, os, time, openpyxl
from selenium import webdriver 
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import timedelta
import os, sys, time, datetime, shutil
# import pillow 

dirpath = ("/Users/thamayron/Dropbox/TELY/SCRIPTS/relatorios/")
dirpath3 = ("/Users/thamayron/Dropbox/TELY/SCRIPTS/")

# MUDA A PASTA DE DOWNLOADS
chromeOptions = webdriver.ChromeOptions()
prefs = {"download.default_directory" : dirpath}
chromeOptions.add_experimental_option("prefs",prefs)

# ENTRAR NO SITE 
pasta_chrome = "/Users/thamayron/Dropbox/TELY/SCRIPTS/chromedriver"
site = 'http://187.33.233.42:8080/queuemetrics/autenticazione.jsp'
browser = webdriver.Chrome(executable_path=pasta_chrome, chrome_options=chromeOptions)
browser.get(site)

# LOGIN E SENHA - CAPTURA OS DADOS DOS CAMPOS
username = browser.find_element_by_id("AUTH_logon")
password = browser.find_element_by_id("AUTH_password")

# LOGIN E SENHA - INSERE OS DADOS NOS CAMPOS
username.send_keys("thamayron")
password.send_keys("123mudar")

# CLICA NO SUBMIT
browser.find_element_by_class_name("actButton").click()

# ENCONTRA A FILA E O RELATÓRIO
fila = browser.find_element_by_id("CODA_input")
relatorio = browser.find_element_by_id("CODA_qmreport")

# SELECINA NOME DA FILA E RELATÓRIO
fila.send_keys("ISP")
relatorio.send_keys("Todos os Relatórios")

# CLICA NO RELATÓRIO ESCOLHIDO
# browser.find_element_by_link_text("Hoje").click()
periodo = "Ontem"
browser.find_element_by_link_text(periodo).click()

# ESCOLHE O FILTRO DE ATENDIDAS
time.sleep(3) # espera 1s antes de fechar
browser.find_element_by_link_text("Atendidas (detalhes)").click()
time.sleep(2) # espera 2s antes de exportar arquivo
# browser.find_element_by_xpath("//input[@value='XLS']").click()
browser.find_element_by_xpath('//*[@title="Selecione as colunas para apresentação"]').click()

todas_opcoes_check = ["OK_tstEnter","OK_from", "OK_queue", "OK_ivrTime", "OK_waitLen", "OK_callLen", "OK_oPos", "OK_reason", "OK_transfer", "OK_agent", "OK_attempts", "OK_stcode", "OK_nStints", "OK_serverId", "OK_callid","OK_mohEvts","OK_mohDur", "OK_ivrWait","OK_ivr","OK_dnis","OK_url", "OK_callTag", "OK_featuresNum", "OK_varsNum", "OK_features","OK_variables"]
opcoes_check = ["OK_tstEnter","OK_from", "OK_waitLen", "OK_callLen", "OK_agent"]

for o in todas_opcoes_check:
	checked = browser.find_element_by_id(o)
	if o in opcoes_check:
		# time.sleep(1) # espera 1s 
		if (checked.is_selected()):
			# print(o+" NAO FACA NADA")
			time.sleep(1) # espera 1s antes de fechar
			pass
		else:
			checked.click()
			print(o+" Checkbox não está selecionado.. agora selecionando")
			time.sleep(1) # espera 1s antes de fechar
	else:
		if (checked.is_selected()):
			checked.click()
			print(o+" Checkbox selecionado.. deselecionando")
			# time.sleep(1) # espera 1s 
		else:
			# time.sleep(1) # espera 1s 
			pass


# time.sleep(1) # espera 1s antes de fechar
time.sleep(2) # espera 2s antes de exportar arquivo
browser.find_element_by_class_name("ui-dialog-buttonset").click()

time.sleep(1) # espera 1s antes de fechar
browser.find_element_by_xpath('//*[@title="Exportar como CSV (Comma-separated values)"]').click()

# ENCERRA O BROWSER
time.sleep(1) # espera 1s antes de fechar
browser.close()

# RENOMEIA O ARQUIVO BAIXADO
arquivos = os.listdir(dirpath) # lista de arquivos da pasta
l = 0
ext = "csv"
for i in arquivos:
	l = l+1 # contador
	ext1 = i.split('.') # separa pelo ponto, vira uma lista
	ext2 = ext1[-1] # mostra só a ultima quebra do nome
	if ext2 == ext:
		nome = "original."+ext
		renomeia = os.rename(dirpath+i,dirpath+nome) # renomeando arquivo

# dia = datetime.datetime.now().isoweekday()
dia = datetime.datetime.now().weekday()
dia = dia-1
if dia == 0:
	dia = "SEG"
elif dia == 1:
	dia = "TER"
elif dia == 2:
	dia = "QUA"
elif dia == 3:
	dia = "QUI"
elif dia == 4:
	dia = "SEX"
elif dia == 5:
	dia = "SAB"
else:
	dia = "DOM"
# if dia == 1:
# 	dia = "SEG"
# elif dia == 2:
# 	dia = "TER"
# elif dia == 3:
# 	dia = "QUA"
# elif dia == 4:
# 	dia = "QUI"
# elif dia == 5:
# 	dia = "SEX"
# elif dia == 6:
# 	dia = "SAB"
# else:
# 	dia = "DOM"


arquivo_origem = (dirpath+nome)
arquivo_destino = (dirpath3+"relatorio.xlsx")
data_ontem = (datetime.datetime.now() - timedelta(1)).strftime("%d_%m_%Y")
nome_folha = dia+"_"+data_ontem
# nome_folha = dia+"_"+datetime.datetime.now().strftime("%d_%m_%Y")
wb2 = load_workbook(arquivo_destino)
ws21 = wb2.active
ws21 = wb2.create_sheet(nome_folha)


file = open(arquivo_origem,"r") 
 
g=0
arquivo=[]
for linha in file:
	if(linha):
		g=g+1	
	linhaarray = linha.split(';')
	Ax = str(linhaarray[0])
	Bx = str(linhaarray[1])
	Cx = str(linhaarray[2])
	Dx = str(linhaarray[3])
	Ex = str(linhaarray[4])
	for c in range(1,6):
		valor = linhaarray[c-1]
		ws21.cell(row=g+1, column=c, value=valor.replace('"',''))
	# print (Ax+' | '+Bx+' | '+Cx+' | '+Dx+' | '+Ex)
file.close() 

wb2[nome_folha].cell(row=1, column=1,value="CHAMADAS ATENDIDAS")
wb2[nome_folha].merge_cells('A1:E1')

# wb2['GERAL'].cell(row=1, column=1, value='DIA DA SEMANA')
wb2['GERAL'].cell(row=1, column=1, value='ATENDIDAS')
wb2['GERAL'].merge_cells('A1:B1')
wb2['GERAL'].cell(row=1, column=8, value='TOTAL DE ATENDIDAS')

k=0
condicao = True
while (condicao):
	k=k+1
	if not ((wb2['GERAL'].cell(row=k+1,column=2).value)):
		condicao = False

wb2['GERAL'].cell(row=k+1, column=1, value=nome_folha)
# wb2['GERAL'].cell(row=k+1, column=2, value='ATENDIDAS')

wb2['GERAL'].cell(row=k+1, column=2, value='=SUM(COUNTA('+nome_folha+'!A:A)-2)')
wb2['GERAL'].cell(row=2, column=8, value='=SUM(B2:B'+str(k+1)+')')

wb2.save(filename = arquivo_destino)

time.sleep(10) # espera 1s antes de fechar
os.remove(dirpath+nome)


"""

# FAZ A EXTRACAO DO ARQUIVO BAIXADO E INSERE NO RELATORIO GERAL
# NOME DOS ARQUIVOS E FOLHAS


# WORKBOOK 1 E 2 (ORIGEM E DESTINO)
wb = load_workbook(arquivo_origem)
wb2 = load_workbook(arquivo_destino)
ws11 = wb.active
ws21 = wb2.active
ws21 = wb2.create_sheet(nome_folha) # CRIA UMA NOVA PLANILHA A CADA EXECUCAO COM DATA E HORA 

# ENQUANTO A CONDICAO FOR VERDADEIRA O LACO E EXECUTADO
condicao = True
k=0
while (condicao):
	k=k+1
	# PEGA O VALOR DE CADA CELULA (LINHA + COLUNA)
	Ax = str(ws11.cell(row=k,column=1).value)
	Bx = str(ws11.cell(row=k,column=2).value)
	Cx = str(ws11.cell(row=k,column=3).value)
	Dx = str(ws11.cell(row=k,column=4).value)
	Ex = str(ws11.cell(row=k,column=5).value)
	# QUANTIDADE DE COLUNAS = 5
	for c in range(1,6):
		valor = str(ws11.cell(row=k,column=c).value)
		ws21.cell(row=k, column=c, value=valor)

	# print (Ax+' | '+Bx+' | '+Cx+' | '+Dx+' | '+Ex)
	if not ((ws11.cell(row=k+1,column=1).value)):
		condicao = False

# CRIA UMA LINHA EM G2 PARA COLOCAR O TOTAL DE LIGACOES
ws21.cell(row=1, column=7, value="TOTAL DE LIGACOES")
ws21.cell(row=2, column=7, value=k-1)

# NA LISTA DE PLANILHAS, VARRE TODAS E PROCURA OS VALORES EM G2, EXCETO NA GERAL
contadorG2 = 0
for sheet in wb2.sheetnames:
	if not sheet == 'GERAL':
		ws11_2 = wb2[sheet].cell(row=2, column=7)
		contadorG2 = int(ws11_2.value) + contadorG2

# ADICIONA O TOTAL DE LIGACOES (EM CADA CELULA G2) A UMA CELULA NA PLANILHA GERAL
ws11_2 = wb2['GERAL'].cell(row=3, column=1,value=contadorG2)

wb.save(filename = arquivo_origem) # SALVA
wb2.save(filename = arquivo_destino) # SALVA

"""
# DATA - OK_tstEnter
# ORIGINADOR - OK_from
# FILA - OK_queue
# URA - OK_ivrTime
# ESPERA - OK_waitLen
# DURACAO - OK_callLen
# POS - OK_oPos
# DESCONEXAO - OK_reason
# TRASNF - OK_transfer
# ATENDIDA POR - OK_agent
# TENTATIVAS - OK_attempts
# CODIGO - OK_stcode
# STINTS - OK_nStints
# SRV - OK_serverId
# ASTERISK - OK_callid
# EVENTOS - OK_mohEvts
# DUR MUS ESP - OK_mohDur
# DUR URA - OK_ivrWait
# CAMINHO URA - OK_ivr
# DNIS - OK_dnis
# URL - OK_url
# MARCACAO - OK_callTag
# FUNC - OK_featuresNum
# VARIAVEIS - OK_varsNum
# FEAT - OK_features
# VARIA - OK_variables

# download googlewebdriver

# https://sites.google.com/a/chromium.org/chromedriver/downloads

# documentação selenium

# http://selenium-python.readthedocs.io/api.html#selenium.common.exceptions.InvalidCookieDomainException

# documentação openyxl

# https://openpyxl.readthedocs.io/en/stable/

# Instalações necessárias

# pip3 install webdriver
# pip3 install openpyxl - excel
# pip3 install pillow - para imgs